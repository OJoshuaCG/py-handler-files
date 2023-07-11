# pip install openpyxl
from openpyxl import Workbook, load_workbook

# Creando un archivo excel y escribiendo sobre el
workbook = Workbook()

sheet = workbook.active
sheet["A1"] = "hola"
sheet["B1"] = "mundo"

workbook.save(filename="archivo.xlsx")
workbook.close()

# Leyendo un archivo excel y explorando sus propiedades
workbook = load_workbook("videogamesales.xlsx")
# workbook = load_workbook("videogamesales.xlsx", read_only=True, data_only=False)

print(workbook.sheetnames)

# Seleccionando la primera hoja (por defecto)
sheet = workbook.active
print(sheet.title)

# Seleccionando una hoja en especifico
# sheet_vgsales = workbook['vgsales']

# print(f"Filas:    {sheet_vgsales.max_row}")
# print(f"Columnas: {sheet_vgsales.max_column}")

# Tomando una celda
celda_A1 = sheet['A1'].value
print(f"Celda A1: {celda_A1}")

celda_A1 = sheet.cell(row=1, column=1).value
print(f"Celda A1: {celda_A1}")

# Obtener un rango de celdas:
data = sheet["A"]       # Columna A
data = sheet["A:B"]     # Columna A -> B
data = sheet["2"]       # Fila 2
data = sheet["2:3"]     # Fila 2 -> 3
data = sheet["A1:C2"]   # Rango A1 -> C2


# Iterar sobre filas:
for row in sheet.iter_rows(min_row=1, max_row=2, min_col=1, max_col=3):
    print(row)

# Iterar sobre columnas:
for column in sheet.iter_rows(min_row=1, max_row=2, min_col=1, max_col=3):
    print(column)

# # Iterar sobre toda la hoja:
# for row in sheet.rows:
#     print(row)

# for column in sheet.colums:
#     print(column)


# Escribiendo en la hoja
sheet['K1'] = "Suma de las ventas"
sheet.cell(row=2, column=11, value="10,000")
workbook.save("copia-videogamesales.xlsx")

workbook.close()

# Source:
# - https://www.datacamp.com/tutorial/python-excel-tutorial
# - https://www.geeksforgeeks.org/working-with-excel-spreadsheets-in-python/
# - https://realpython.com/openpyxl-excel-spreadsheets-python/
# - https://pybonacci.org/2021/03/10/curso-sobre-como-trabajar-con-hojas-de-calculo-excel-calc-usando-openpyxl-en-python-iii/

# # Iterando sobre una fila, en este caso, el encabezado
# header = [sheet.cell(
#     row=1, column=i).value for i in range(1, sheet.max_column+1)]
# print(header)

# # Iterando sobre una columna
# data_B = [sheet.cell(row=i, column=2).value for i in range(2, 12)]
# print(data_B)
